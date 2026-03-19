"""Report generation service (B10.1, B10.2, B10.3).

Generates reports synchronously on request (Phase 1 simplification).
For large datasets, this can be moved to a background task in Phase 2.
"""

import csv
import io
import uuid
from datetime import UTC, datetime

from sqlalchemy.ext.asyncio import AsyncSession

from app.dashboards import service as dashboard_svc
from app.reports.models import ReportRequest
from app.reports.schemas import VALID_REPORT_TYPES
from app.shared.exceptions import NotFoundException, ValidationException


async def list_reports(
    db: AsyncSession,
    *,
    requested_by: uuid.UUID,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[ReportRequest], int]:
    """List reports for a user, newest first."""
    from sqlalchemy import func, select

    base = select(ReportRequest).where(ReportRequest.requested_by == requested_by)
    total_q = select(func.count()).select_from(base.subquery())
    total = (await db.execute(total_q)).scalar() or 0

    rows_q = (
        base.order_by(ReportRequest.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    rows = (await db.execute(rows_q)).scalars().all()
    return list(rows), total


async def create_report(
    db: AsyncSession,
    *,
    requested_by: uuid.UUID,
    report_type: str,
    fmt: str,
    filters: dict | None = None,
) -> ReportRequest:
    """Create a report request and generate the data immediately."""
    if report_type not in VALID_REPORT_TYPES:
        raise ValidationException(
            f"Invalid report_type: {report_type}. Valid: {', '.join(VALID_REPORT_TYPES)}"
        )

    report = ReportRequest(
        requested_by=requested_by,
        report_type=report_type,
        format=fmt,
        filters=filters or {},
        status="pending",
    )
    db.add(report)
    await db.flush()

    try:
        await _generate_report(db, report)
        report.status = "completed"
        report.completed_at = datetime.now(UTC)
    except Exception as e:
        report.status = "failed"
        report.error_message = str(e)[:500]

    await db.commit()
    await db.refresh(report)
    return report


async def get_report(db: AsyncSession, report_id: uuid.UUID) -> ReportRequest:
    """Get a report by ID."""
    report = await db.get(ReportRequest, report_id)
    if not report:
        raise NotFoundException("Report not found")
    return report


async def generate_download(db: AsyncSession, report_id: uuid.UUID) -> tuple[bytes, str, str]:
    """Generate and return downloadable content.

    Returns: (content_bytes, content_type, filename)
    """
    report = await get_report(db, report_id)
    if report.status != "completed":
        raise ValidationException(f"Report is not ready (status: {report.status})")

    data = await _fetch_report_data(db, report)

    if report.format == "csv":
        content, content_type = _to_csv(data, report.report_type)
        filename = f"{report.report_type}_{report.id}.csv"
    elif report.format == "xlsx":
        content, content_type = await _to_xlsx(data, report.report_type)
        filename = f"{report.report_type}_{report.id}.xlsx"
    elif report.format == "pdf":
        content, content_type = _to_pdf(data, report.report_type)
        filename = f"{report.report_type}_{report.id}.pdf"
    else:
        content, content_type = _to_csv(data, report.report_type)
        filename = f"{report.report_type}_{report.id}.csv"

    return content, content_type, filename


# ═══════════════════════════════════════════════════════════════
# Internal helpers
# ═══════════════════════════════════════════════════════════════

async def _generate_report(db: AsyncSession, report: ReportRequest) -> None:
    """Validate that the report data can be generated (pre-check)."""
    await _fetch_report_data(db, report)


async def _fetch_report_data(db: AsyncSession, report: ReportRequest) -> list[dict]:
    """Fetch the underlying data for a report type."""
    rt = report.report_type
    filters = report.filters or {}

    if rt == "monthly_biosecurity_summary":
        return [await dashboard_svc.executive_summary(db)]

    if rt == "farm_score_comparison":
        return await dashboard_svc.benchmark(
            db,
            farm_type=filters.get("farm_type"),
            region_id=filters.get("region_id"),
        )

    if rt == "trust_gap_report":
        return await dashboard_svc.trust_gaps(db)

    if rt == "killer_metrics_summary":
        return await dashboard_svc.killer_metrics_trend(
            db,
            farm_id=filters.get("farm_id"),
            months=filters.get("months", 6),
        )

    if rt == "scar_hotspot_report":
        return await dashboard_svc.scar_hotspots(
            db,
            farm_id=filters.get("farm_id"),
        )

    if rt in ("case_backlog", "overdue_tasks"):
        # Use executive summary as a summary source
        summary = await dashboard_svc.executive_summary(db)
        return [summary]

    return []


def _to_csv(data: list[dict], report_type: str) -> tuple[bytes, str]:
    """Convert list of dicts to CSV bytes."""
    if not data:
        return b"", "text/csv"

    output = io.StringIO()
    # Flatten nested dicts for CSV
    flat_data = _flatten_rows(data)
    if not flat_data:
        return b"", "text/csv"

    writer = csv.DictWriter(output, fieldnames=flat_data[0].keys())
    writer.writeheader()
    writer.writerows(flat_data)
    return output.getvalue().encode("utf-8-sig"), "text/csv"


async def _to_xlsx(data: list[dict], report_type: str) -> tuple[bytes, str]:
    """Convert list of dicts to XLSX bytes using openpyxl."""
    try:
        from openpyxl import Workbook
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        content, _ = _to_csv(data, report_type)
        return content, "text/csv"

    wb = Workbook()
    ws = wb.active
    ws.title = report_type[:31]  # Excel sheet name max 31 chars

    flat_data = _flatten_rows(data)
    if not flat_data:
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Header row
    headers = list(flat_data[0].keys())
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for row in flat_data:
        ws.append([row.get(h) for h in headers])

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        max_len = max(len(str(header)), *(len(str(row.get(header, ""))) for row in flat_data))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _to_pdf(data: list[dict], report_type: str) -> tuple[bytes, str]:
    """Convert list of dicts to PDF bytes using fpdf2."""
    from fpdf import FPDF

    _DEJAVU = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    _DEJAVU_BOLD = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"

    flat_data = _flatten_rows(data)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_font("DejaVu", "", _DEJAVU)
    pdf.add_font("DejaVu", "B", _DEJAVU_BOLD)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title
    pdf.set_font("DejaVu", "B", 14)
    title = report_type.replace("_", " ").title()
    pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)

    if not flat_data:
        pdf.set_font("DejaVu", "", 11)
        pdf.cell(0, 10, "No data available.", new_x="LMARGIN", new_y="NEXT", align="C")
        output = io.BytesIO()
        pdf.output(output)
        return output.getvalue(), "application/pdf"

    headers = list(flat_data[0].keys())
    num_cols = len(headers)
    page_width = pdf.w - pdf.l_margin - pdf.r_margin
    col_width = page_width / num_cols if num_cols else page_width

    # Header row
    pdf.set_font("DejaVu", "B", 8)
    pdf.set_fill_color(220, 220, 220)
    for h in headers:
        pdf.cell(col_width, 7, str(h)[:20], border=1, fill=True, align="C")
    pdf.ln()

    # Data rows
    pdf.set_font("DejaVu", "", 7)
    for row in flat_data:
        for h in headers:
            val = str(row.get(h, ""))[:25]
            pdf.cell(col_width, 6, val, border=1)
        pdf.ln()

    output = io.BytesIO()
    pdf.output(output)
    return output.getvalue(), "application/pdf"


def _flatten_rows(data: list[dict]) -> list[dict]:
    """Flatten nested dicts/lists for tabular export."""
    flat = []
    for row in data:
        flat_row = {}
        for k, v in row.items():
            if isinstance(v, dict):
                for sk, sv in v.items():
                    flat_row[f"{k}_{sk}"] = sv
            elif isinstance(v, list):
                flat_row[k] = len(v)  # Just count for lists
            else:
                flat_row[k] = v
        flat.append(flat_row)
    return flat
